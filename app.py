from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, jsonify, flash)
import mysql.connector
import os, io, re
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import landscape, A5, A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
 
app = Flask(__name__)
app.secret_key = 'sat2026_csc_secret_xK9mP3nQ'
 
# ─── DATABASE CONFIG ─────────────────────────────────────────
DB_CONFIG = {
    'host':     'sql12.freesqldatabase.com',
    'user':     'sql12820158',
    'password': 'sRMDyVMtMs',
    'database': 'sql12820158',
    'port':     3306,
    'charset':  'utf8mb4',
}
 
# ─── ADMIN CREDENTIALS ───────────────────────────────────────
ADMIN_USERNAME = 'admin'
ADMIN_PASSWORD = 'sat@admin2026'
 
# ─── EXAM CONFIGURATION ──────────────────────────────────────
EXAM_DATES = [
    '29 March 2026 (Sunday)',
    '04 April 2026 (Saturday)',
]
 
TIMINGS = [
    '7:00 AM', '7:30 AM',
    '8:00 AM', '8:30 AM',
    '9:00 AM', '9:30 AM',
    '10:00 AM','10:30 AM',
    '11:00 AM','11:30 AM',
    '12:00 PM','12:30 PM',
    '1:00 PM',
]
 
# prefix per centre — used in hall ticket no generation
CENTRES = {
    'Pammal':     'pmld',
    'Pallavaram': 'pvmd',
    'Chrompet':   'chrd',
}
 
# ─── DATABASE HELPERS ────────────────────────────────────────
def get_db():
    return mysql.connector.connect(**DB_CONFIG)
 
 
def init_db():
    """Create the registrations table if it doesn't exist."""
    try:
        conn = get_db()
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"  ❌  Cannot connect to database")
        print(f"  Error: {e}")
        print(f"{'='*60}\n")
        raise SystemExit(1)
 
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS registrations (
            id            INT AUTO_INCREMENT PRIMARY KEY,
            admit_card_no VARCHAR(20)  UNIQUE NOT NULL,
            name          VARCHAR(100) NOT NULL,
            gender        ENUM('Male','Female') NOT NULL,
            mobile        VARCHAR(15)  NOT NULL,
            exam_centre   VARCHAR(30)  NOT NULL,
            exam_date     VARCHAR(50)  NOT NULL,
            exam_time     VARCHAR(15)  NOT NULL,
            registered_at TIMESTAMP    DEFAULT CURRENT_TIMESTAMP,
            is_active     TINYINT(1)   DEFAULT 1,
            INDEX idx_mobile (mobile),
            INDEX idx_admit  (admit_card_no),
            INDEX idx_centre (exam_centre)
        )
    """)
    conn.commit()
    cur.close()
    conn.close()
    print("✅  Database tables ready.")
 
 
def generate_admit_no(centre):
    """Generate hall ticket: prefix + zero-padded serial per centre."""
    prefix = CENTRES.get(centre, 'sat')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM registrations WHERE exam_centre=%s", (centre,)
    )
    count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return f"{prefix}{count + 1:03d}"
 
 
# ─── AUTH DECORATOR ──────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated
 
 
# ─── ADMIT CARD PDF GENERATOR ────────────────────────────────
def build_admit_pdf(s):
    """
    Generates a pixel-perfect CSC SAT 2026 admit card as a PDF in memory.
    Matches the original printed design:
      - Landscape A5
      - Left panel: red bg, SAT2026 badge, 75% scholarship box,
        Tamil text, CSC circle logo
      - Decorative gold+green diamond border strips both sides
      - Right panel: cream bg, ADMIT CARD title, 6-box admit no,
        Sex checkboxes, Time field, Name underline,
        Centre Address, desk-calendar widget, Examiner
    """
    buf = io.BytesIO()
    W, H = landscape(A5)          # 595.3 × 419.5 pt
    c    = pdfcanvas.Canvas(buf, pagesize=(W, H))
 
    # Colours (sampled from reference image)
    RED        = colors.HexColor('#FF0000')
    CREAM      = colors.HexColor('#FCFDCD')
    NAVY       = colors.HexColor('#0000FF')
    DARK_NAVY  = colors.HexColor('#0808A6')
    YELLOW     = colors.HexColor('#E4FF00')
    WHITE      = colors.white
    GOLD       = colors.HexColor('#C8A050')
    GREEN_BORD = colors.HexColor('#A8D898')
 
    LP = 58 * mm    # left panel width
    BR = 13 * mm    # decorative border strip width (each side)
 
    # ── Cream background ──────────────────────────────────────
    c.setFillColor(CREAM)
    c.rect(0, 0, W, H, fill=1, stroke=0)
 
    # ── Decorative border strips ──────────────────────────────
    def draw_border(x, sw, h):
        c.setFillColor(GOLD);      c.rect(x, 0, sw, h, fill=1, stroke=0)
        c.setFillColor(CREAM);     c.rect(x+2, 2, sw-4, h-4, fill=1, stroke=0)
        c.setFillColor(GREEN_BORD);c.rect(x+4, 4, sw-8, h-8, fill=1, stroke=0)
        c.setFillColor(CREAM);     c.rect(x+7, 7, sw-14, h-14, fill=1, stroke=0)
        cx = x + sw/2
        c.setFillColor(GOLD)
        for yp in range(10, int(h), 10):
            p = c.beginPath()
            p.moveTo(cx, yp+4); p.lineTo(cx+4, yp)
            p.lineTo(cx, yp-4); p.lineTo(cx-4, yp); p.close()
            c.drawPath(p, fill=1, stroke=0)
        c.setFillColor(GREEN_BORD)
        for yp in range(5, int(h), 10):
            c.circle(cx, yp, 1.5, fill=1, stroke=0)
 
    draw_border(0, BR, H)
    draw_border(W - BR, BR, H)
 
    # ── Left red panel ────────────────────────────────────────
    c.setFillColor(RED)
    c.rect(BR, 0, LP - BR, H, fill=1, stroke=0)
    mid = BR + (LP - BR) / 2     # horizontal centre of left panel
 
    # SAT 2026 badge
    bx = BR+6; by = H-52; bw = LP-BR-12; bh = 40
    c.setFillColor(DARK_NAVY); c.roundRect(bx, by, bw, bh, 10, fill=1, stroke=0)
    c.setStrokeColor(YELLOW); c.setLineWidth(2.5)
    c.ellipse(bx+4, by+4, bx+bw-4, by+bh-4, fill=0, stroke=1)
    c.setFillColor(YELLOW); c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(mid, by+bh-15, 'SAT 2026')
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 5.5)
    c.drawCentredString(mid, by+5, '33rd Scholarship Aptitude Test')
 
    # Win up to 75% box
    wx = BR+5; wy = H-148; ww = LP-BR-10; wh = 90
    c.setFillColor(DARK_NAVY); c.roundRect(wx, wy, ww, wh, 8, fill=1, stroke=0)
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(mid, wy+wh-14, 'Win up to')
    c.setFillColor(YELLOW); c.setFont('Helvetica-Bold', 46)
    c.drawCentredString(mid, wy+wh-60, '75%')
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(mid, wy+8, 'Scholarship')
 
    # Tamil text
    c.setFillColor(DARK_NAVY); c.setFont('Helvetica-Bold', 7)
    c.drawCentredString(mid, H-162, 'anumadhi ilavasam !')
    c.drawCentredString(mid, H-173, 'anaivarium varuga !!')
 
    # CSC circle logo
    cy0 = 34
    c.setFillColor(DARK_NAVY); c.circle(mid, cy0, 28, fill=1, stroke=0)
    c.setStrokeColor(YELLOW);  c.setLineWidth(2.5)
    c.circle(mid, cy0, 28, fill=0, stroke=1)
    c.setFillColor(YELLOW); c.setFont('Helvetica-BoldOblique', 26)
    c.drawCentredString(mid, cy0-9, 'CSC')
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 4.5)
    c.drawCentredString(mid, cy0-19, 'COMPUTER SOFTWARE COLLEGE')
    c.drawCentredString(mid, cy0-25, 'An ISO 9001 : 2015 Certified Institution')
 
    # ── Right cream panel ─────────────────────────────────────
    rx  = LP + 10                       # left margin of right content
    rc  = LP + (W - BR - LP) / 2       # horizontal centre of right panel
 
    # "ADMIT CARD" title
    c.setFillColor(RED); c.setFont('Helvetica-Bold', 38)
    c.drawCentredString(rc, H-40, 'ADMIT CARD')
 
    # Admit Card No label + 6 boxes
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(rx, H-102, 'Admit Card No:')
    bx0  = rx+80; bry = H-125; bw2 = 32; bh2 = 30; bgap = 4
    admit_no = str(s.get('admit_card_no', ''))
    for i in range(6):
        xi = bx0 + i*(bw2+bgap)
        c.setFillColor(CREAM); c.setStrokeColor(NAVY); c.setLineWidth(1.5)
        c.rect(xi, bry, bw2, bh2, fill=1, stroke=1)
        if i < len(admit_no):
            c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
            c.drawCentredString(xi+bw2/2, bry+9, admit_no[i].upper())
 
    # Sex + Time row
    sy = H-177
    ms = 15    # checkbox size
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(rx, sy, 'Sex:')
    mx = rx+32; my = sy-11
    c.setFillColor(CREAM); c.setStrokeColor(NAVY); c.setLineWidth(1.5)
    c.rect(mx, my, ms, ms, fill=1, stroke=1)
    if s.get('gender') == 'Male':
        c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 10)
        c.drawCentredString(mx+ms/2, my+3, 'X')
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(mx+ms+4, sy, 'M')
    fx = mx+ms+22
    c.setFillColor(CREAM); c.setStrokeColor(NAVY); c.setLineWidth(1.5)
    c.rect(fx, my, ms, ms, fill=1, stroke=1)
    if s.get('gender') == 'Female':
        c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 10)
        c.drawCentredString(fx+ms/2, my+3, 'X')
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(fx+ms+4, sy, 'F')
    tlx = fx+ms+30
    c.drawString(tlx, sy, 'Time:')
    tvx = tlx+38
    time_val = str(s.get('exam_time', ''))
    c.drawString(tvx, sy, time_val)
    c.setStrokeColor(NAVY); c.setLineWidth(1.2)
    c.line(tvx, sy-3, tvx+72, sy-3)
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(tvx+74, sy, 'am./pm.')
 
    # Name row
    ny = H-210
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(rx, ny, 'Name :')
    c.drawString(rx+50, ny, str(s.get('name', '')).upper())
    c.setStrokeColor(NAVY); c.setLineWidth(1.2)
    c.line(rx+48, ny-3, W-BR-14, ny-3)
 
    # Centre Address row
    cay = H-243
    c.setFillColor(NAVY); c.setFont('Helvetica-Bold', 11)
    c.drawString(rx, cay, 'Centre Address :')
    c.drawString(rx+104, cay, str(s.get('exam_centre', '')).upper())
 
    # Desk calendar widget
    exam_date_str = str(s.get('exam_date', ''))
    if 'April' in exam_date_str:
        cal_month, cal_day, cal_day_name = 'APRIL', '4', 'SATURDAY'
    else:
        cal_month, cal_day, cal_day_name = 'MARCH', '29', 'SUNDAY'
 
    clx = LP+12; cly = H-344; clw = 68; clh = 90
    c.setFillColor(colors.HexColor('#999999'))
    c.roundRect(clx+3, cly-3, clw, clh, 4, fill=1, stroke=0)   # shadow
    c.setFillColor(DARK_NAVY)
    c.roundRect(clx, cly, clw, clh, 4, fill=1, stroke=0)
    th = 22
    c.setFillColor(RED)
    c.roundRect(clx, cly+clh-th, clw, th, 4, fill=1, stroke=0)
    c.rect(clx, cly+clh-th, clw, th/2, fill=1, stroke=0)
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 7)
    c.drawCentredString(clx+clw/2, cly+clh-10, 'EXAM DATE')
    c.setFillColor(YELLOW); c.setFont('Helvetica-Bold', 14)
    c.drawCentredString(clx+clw/2, cly+clh-38, cal_month)
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 36)
    c.drawCentredString(clx+clw/2, cly+clh-72, cal_day)
    c.setFillColor(RED)
    c.roundRect(clx+5, cly+3, clw-10, 14, 3, fill=1, stroke=0)
    c.setFillColor(WHITE); c.setFont('Helvetica-Bold', 7)
    c.drawCentredString(clx+clw/2, cly+6, cal_day_name)
 
    # "Examiner" bottom-right
    c.setFillColor(RED); c.setFont('Helvetica-Bold', 11)
    c.drawString(W-BR-58, H-386, 'Examiner')
 
    c.save()
    buf.seek(0)
    return buf
 
 
# ═══════════════════════════════════════════════════════════
# PUBLIC ROUTES
# ═══════════════════════════════════════════════════════════
 
@app.route('/')
def index():
    return render_template('index.html')
 
 
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name   = request.form.get('name',        '').strip()
        gender = request.form.get('gender',      '')
        mobile = request.form.get('mobile',      '').strip()
        centre = request.form.get('exam_centre', '')
        date   = request.form.get('exam_date',   '')
        time   = request.form.get('exam_time',   '')
 
        # Validate
        errors = []
        if not name:
            errors.append('Name is required.')
        if gender not in ('Male', 'Female'):
            errors.append('Please select your gender.')
        if not re.match(r'^\d{10}$', mobile):
            errors.append('Enter a valid 10-digit mobile number.')
        if centre not in CENTRES:
            errors.append('Please select a valid exam centre.')
        if date not in EXAM_DATES:
            errors.append('Please select a valid exam date.')
        if time not in TIMINGS:
            errors.append('Please select a valid time slot.')
 
        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('register.html', form=request.form,
                                   centres=CENTRES, timings=TIMINGS,
                                   exam_dates=EXAM_DATES)
 
        try:
            conn     = get_db()
            cur      = conn.cursor()
            admit_no = generate_admit_no(centre)
            cur.execute("""
                INSERT INTO registrations
                  (admit_card_no, name, gender, mobile,
                   exam_centre,   exam_date, exam_time)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (admit_no, name, gender, mobile, centre, date, time))
            conn.commit()
            cur.close()
            conn.close()
            flash(
                f'Registration successful! '
                f'Your Hall Ticket No: <strong>{admit_no}</strong>. '
                f'Enter your mobile number to download your admit card.',
                'success'
            )
            return redirect(url_for('check_admit'))
        except Exception as e:
            flash(f'Registration failed: {str(e)}', 'danger')
 
    return render_template('register.html', form={},
                           centres=CENTRES, timings=TIMINGS,
                           exam_dates=EXAM_DATES)
 
 
@app.route('/check', methods=['GET', 'POST'])
def check_admit():
    student  = None
    students = []
    error    = None
 
    if request.method == 'POST':
        mobile = request.form.get('mobile', '').strip()
        if not mobile:
            error = 'Please enter your mobile number.'
        else:
            try:
                conn = get_db()
                cur  = conn.cursor(dictionary=True)
                cur.execute(
                    "SELECT * FROM registrations "
                    "WHERE mobile=%s AND is_active=1 "
                    "ORDER BY registered_at DESC",
                    (mobile,)
                )
                students = cur.fetchall()
                cur.close()
                conn.close()
                if not students:
                    error = 'No registration found for this mobile number.'
                elif len(students) == 1:
                    student = students[0]
            except Exception as e:
                error = f'Database error: {str(e)}'
 
    return render_template('check_admit.html',
                           student=student, students=students, error=error)
 
 
@app.route('/download_admit/<int:reg_id>')
def download_admit(reg_id):
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT * FROM registrations WHERE id=%s AND is_active=1",
            (reg_id,)
        )
        s = cur.fetchone()
        cur.close()
        conn.close()
        if not s:
            flash('Record not found.', 'danger')
            return redirect(url_for('check_admit'))
        buf   = build_admit_pdf(s)
        fname = f"AdmitCard_{s['admit_card_no']}.pdf"
        return send_file(buf, as_attachment=True,
                         download_name=fname, mimetype='application/pdf')
    except Exception as e:
        flash(f'Error generating admit card: {str(e)}', 'danger')
        return redirect(url_for('check_admit'))
 
 
# ═══════════════════════════════════════════════════════════
# ADMIN ROUTES
# ═══════════════════════════════════════════════════════════
 
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if (request.form.get('username') == ADMIN_USERNAME and
                request.form.get('password') == ADMIN_PASSWORD):
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        flash('Invalid credentials.', 'danger')
    return render_template('admin_login.html')
 
 
@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))
 
 
@app.route('/admin')
@admin_required
def admin_dashboard():
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute("SELECT COUNT(*) AS total FROM registrations")
        total = cur.fetchone()['total']
        cur.execute(
            "SELECT exam_centre, COUNT(*) AS cnt "
            "FROM registrations GROUP BY exam_centre"
        )
        by_centre = cur.fetchall()
        cur.execute(
            "SELECT gender, COUNT(*) AS cnt "
            "FROM registrations GROUP BY gender"
        )
        by_gender = cur.fetchall()
        cur.execute(
            "SELECT exam_date, COUNT(*) AS cnt "
            "FROM registrations GROUP BY exam_date"
        )
        by_date = cur.fetchall()
        cur.execute(
            "SELECT * FROM registrations "
            "ORDER BY registered_at DESC LIMIT 10"
        )
        recent = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('admin_dashboard.html',
                               total=total, by_centre=by_centre,
                               by_gender=by_gender, by_date=by_date,
                               recent=recent)
    except Exception as e:
        flash(f'Database error: {str(e)}', 'danger')
        return render_template('admin_dashboard.html',
                               total=0, by_centre=[], by_gender=[],
                               by_date=[], recent=[])
 
 
@app.route('/admin/students')
@admin_required
def admin_students():
    search = request.args.get('q',      '').strip()
    centre = request.args.get('centre', '').strip()
    date   = request.args.get('date',   '').strip()
    gender = request.args.get('gender', '').strip()
    page   = int(request.args.get('page', 1))
    per    = 20
 
    try:
        conn   = get_db()
        cur    = conn.cursor(dictionary=True)
        where  = ['is_active=1']
        params = []
 
        if search:
            where.append(
                '(name LIKE %s OR mobile LIKE %s OR admit_card_no LIKE %s)'
            )
            params += [f'%{search}%'] * 3
        if centre:
            where.append('exam_centre=%s'); params.append(centre)
        if date:
            where.append('exam_date=%s');   params.append(date)
        if gender:
            where.append('gender=%s');      params.append(gender)
 
        wsql = 'WHERE ' + ' AND '.join(where)
        cur.execute(
            f"SELECT COUNT(*) AS cnt FROM registrations {wsql}", params
        )
        total_rows = cur.fetchone()['cnt']
        cur.execute(
            f"SELECT * FROM registrations {wsql} "
            f"ORDER BY registered_at DESC LIMIT %s OFFSET %s",
            params + [per, (page - 1) * per]
        )
        students = cur.fetchall()
        cur.close()
        conn.close()
 
        return render_template(
            'admin_students.html',
            students=students, search=search, centre=centre,
            date=date, gender=gender, page=page,
            total_pages=(total_rows + per - 1) // per,
            total_rows=total_rows,
            centres=list(CENTRES.keys()),
            exam_dates=EXAM_DATES,
        )
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template(
            'admin_students.html',
            students=[], search='', centre='', date='', gender='',
            page=1, total_pages=1, total_rows=0,
            centres=list(CENTRES.keys()), exam_dates=EXAM_DATES,
        )
 
 
@app.route('/admin/download_admit/<int:reg_id>')
@admin_required
def admin_download_admit(reg_id):
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM registrations WHERE id=%s", (reg_id,))
        s = cur.fetchone()
        cur.close()
        conn.close()
        if not s:
            flash('Record not found.', 'danger')
            return redirect(url_for('admin_students'))
        buf = build_admit_pdf(s)
        return send_file(buf, as_attachment=True,
                         download_name=f"AdmitCard_{s['admit_card_no']}.pdf",
                         mimetype='application/pdf')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin_students'))
 
 
@app.route('/admin/download_excel')
@admin_required
def download_excel():
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT admit_card_no, name, gender, mobile, exam_centre, "
            "exam_date, exam_time, registered_at "
            "FROM registrations WHERE is_active=1 "
            "ORDER BY exam_centre, exam_date, exam_time"
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SAT 2026 Registrations'
 
        thin = Side(style='thin')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdrs = ['Hall Ticket No', 'Name', 'Gender', 'Mobile',
                'Centre', 'Exam Date', 'Time', 'Registered At']
 
        for col, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color='FFFFFF', size=11)
            cell.fill      = PatternFill('solid', fgColor='1A237E')
            cell.alignment = Alignment(horizontal='center')
            cell.border    = bdr
 
        for ri, r in enumerate(rows, 2):
            vals = [
                r['admit_card_no'], r['name'],      r['gender'],
                r['mobile'],        r['exam_centre'],r['exam_date'],
                r['exam_time'],     str(r.get('registered_at', ''))
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = bdr
                if ri % 2 == 0:
                    cell.fill = PatternFill('solid', fgColor='E8EAF6')
 
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(cell.value or '')) for cell in col) + 4, 40
            )
 
        buf   = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"SAT2026_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Excel export failed: {str(e)}', 'danger')
        return redirect(url_for('admin_students'))
 
 
@app.route('/admin/download_pdf_list')
@admin_required
def download_pdf_list():
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT admit_card_no, name, gender, mobile, exam_centre, "
            "exam_date, exam_time, registered_at "
            "FROM registrations WHERE is_active=1 "
            "ORDER BY exam_centre, exam_date, exam_time"
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
 
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm
        )
        title_s = ParagraphStyle(
            't', fontSize=16, fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1A237E'), spaceAfter=4
        )
        sub_s = ParagraphStyle(
            's', fontSize=9, alignment=TA_CENTER,
            textColor=colors.gray, spaceAfter=12
        )
        elems = [
            Paragraph('CSC SAT 2026 — Registration List', title_s),
            Paragraph(
                f'Generated: {datetime.now().strftime("%d %B %Y %I:%M %p")}'
                f'  |  Total: {len(rows)}',
                sub_s
            ),
        ]
 
        td = [['#', 'Hall Ticket', 'Name', 'Gender', 'Mobile',
               'Centre', 'Date', 'Time', 'Registered']]
        for i, r in enumerate(rows, 1):
            ra = r.get('registered_at', '')
            if hasattr(ra, 'strftime'):
                ra = ra.strftime('%d/%m/%Y %H:%M')
            td.append([
                str(i), r['admit_card_no'], r['name'], r['gender'],
                r['mobile'], r['exam_centre'], r['exam_date'],
                r['exam_time'], str(ra)
            ])
 
        t = Table(
            td, repeatRows=1,
            colWidths=[10*mm,28*mm,50*mm,20*mm,30*mm,
                       28*mm,48*mm,22*mm,38*mm]
        )
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR',     (0,0),(-1,0), colors.white),
            ('FONTNAME',      (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0),(-1,-1),8),
            ('ALIGN',         (0,0),(-1,-1),'CENTER'),
            ('VALIGN',        (0,0),(-1,-1),'MIDDLE'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),
             [colors.white, colors.HexColor('#E8EAF6')]),
            ('GRID',          (0,0),(-1,-1),
             0.5, colors.HexColor('#C5CAE9')),
            ('TOPPADDING',    (0,0),(-1,-1),4),
            ('BOTTOMPADDING', (0,0),(-1,-1),4),
        ]))
        elems.append(t)
        doc.build(elems)
        buf.seek(0)
        fname = f"SAT2026_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return send_file(buf, as_attachment=True,
                         download_name=fname, mimetype='application/pdf')
    except Exception as e:
        flash(f'PDF export failed: {str(e)}', 'danger')
        return redirect(url_for('admin_students'))
 
 
@app.route('/admin/api/stats')
@admin_required
def api_stats():
    try:
        conn = get_db()
        cur  = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT exam_centre, COUNT(*) AS cnt "
            "FROM registrations GROUP BY exam_centre"
        )
        by_centre = {r['exam_centre']: r['cnt'] for r in cur.fetchall()}
        cur.execute(
            "SELECT gender, COUNT(*) AS cnt "
            "FROM registrations GROUP BY gender"
        )
        by_gender = {r['gender']: r['cnt'] for r in cur.fetchall()}
        cur.close()
        conn.close()
        return jsonify({'by_centre': by_centre, 'by_gender': by_gender})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
 
 
# ═══════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════
 
if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
